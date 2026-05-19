from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

BLUE_DARK   = "1E3A5F"
WHITE       = "FFFFFF"
RED_FILL    = "FF4444"
ORANGE_FILL = "FF8C00"
YELLOW_FILL = "FFD700"
GREEN_FILL  = "28A745"
GRAY_LIGHT  = "F5F5F5"

def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

thin = Side(style="thin", color="CCCCCC")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── SHEET 1 ───────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Bug Report"
ws.sheet_view.showGridLines = False

headers    = ["Bug ID","Title","Severity","Module","Steps to Reproduce",
              "Expected Behavior","Actual Behavior","Browser / Device","Evidence","Status","Date Found"]
col_widths = [10, 42, 12, 22, 52, 40, 44, 24, 22, 10, 14]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill      = solid(BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = brd
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[1].height = 32

bugs = [
    # ── CRITICAL ─────────────────────────────────────────────────────────────
    ("BUG-001",
     "Public enrollment form crashes with accented characters in student name",
     "Critical", "Matricula Publica",
     "1. Go to /matricula.\n2. Enter student name with accent or apostrophe (e.g. 'Maria Jose O'Brien').\n3. Click Continuar.",
     "Form saves correctly and proceeds to payment step.",
     "Server returns HTTP 500. Enrollment record is not created. Student cannot continue registration.",
     "Chrome 124 / Windows 11", "Screenshot attached", "Open", "2025-02-10"),

    ("BUG-002",
     "Wompi webhook does not update enrollment status after successful payment",
     "Critical", "Pagos / Wompi",
     "1. Complete enrollment at /matricula.\n2. Proceed to Wompi checkout.\n3. Pay successfully.\n4. Check enrollment status in admin panel.",
     "Enrollment status changes from 'waiting' to 'active' automatically after payment confirmation.",
     "Enrollment remains in 'waiting' status indefinitely. Webhook endpoint /webhook/wompi does not process the event. Manual intervention required.",
     "Wompi Sandbox / Chrome 124", "Loom link attached", "Open", "2025-02-14"),

    ("BUG-003",
     "Password reset link is invalid immediately after it is sent",
     "Critical", "Autenticacion",
     "1. Go to login page.\n2. Click 'Forgot password'.\n3. Enter registered email.\n4. Open the email immediately.\n5. Click the reset link.",
     "Link is valid for at least 60 minutes and allows the user to set a new password.",
     "Link returns 'Este enlace ha expirado o no es valido' immediately. User cannot recover access to their account.",
     "Gmail / Firefox 125 / macOS", "Screenshot attached", "Open", "2025-02-18"),

    # ── MAJOR ────────────────────────────────────────────────────────────────
    ("BUG-004",
     "Attendance date range filter ignores the end date in Reports",
     "Major", "Reportes de Pagos",
     "1. Go to /reportes/pagos.\n2. Set date range: 2025-01-01 to 2025-01-31.\n3. Click Aplicar.",
     "Report shows only transactions within the selected date range.",
     "Report returns all transactions from 2025-01-01 to present. End date parameter is ignored by the backend query.",
     "Chrome 124 / Windows 11", "Screenshot attached", "Open", "2025-02-20"),

    ("BUG-005",
     "Student role assignment reverts to default after saving in User Management",
     "Major", "Gestion de Usuarios",
     "1. Go to /seguridad/usuarios.\n2. Open a student record.\n3. Change role to 'Profesor'.\n4. Save.\n5. Refresh the page.",
     "Role change persists and is reflected immediately after save.",
     "Role reverts to 'Estudiante' after page refresh. Change is not persisted to the database. No error is displayed.",
     "Safari 17 / macOS", "Loom link attached", "Open", "2025-03-03"),

    ("BUG-006",
     "Export CSV in Attendance Control downloads an empty file when filters are active",
     "Major", "Control de Asistencia",
     "1. Go to /asistencia.\n2. Filter by program or date.\n3. Click Exportar.",
     "CSV file contains only the rows matching the active filters.",
     "Downloaded CSV is empty (headers only, 0 data rows). The export function ignores all active filters.",
     "Firefox 125 / Windows 11", "Loom link attached", "Open", "2025-03-07"),

    ("BUG-007",
     "Teacher portal sidebar overlaps content on mobile screens",
     "Major", "Portal Profesor",
     "1. Open /profesor/mis-grupos on a mobile device (375px width).\n2. Tap the hamburger menu icon.",
     "Sidebar slides in and page content shifts or remains accessible.",
     "Sidebar overlaps the entire content area. Group cards are not clickable. Close button is partially hidden.",
     "Chrome Mobile / iPhone 14", "Screenshot attached", "Open", "2025-03-12"),

    ("BUG-008",
     "Deleting an enrollment has no confirmation dialog",
     "Major", "Gestion de Usuarios",
     "1. Go to /seguridad/usuarios.\n2. Open any student with an active enrollment.\n3. Click the delete enrollment button.",
     "A confirmation dialog appears before permanently cancelling the enrollment.",
     "Enrollment is cancelled immediately with no warning. Action cannot be undone. No success or error toast is shown.",
     "Chrome 124 / Windows 11", "Screenshot attached", "Open", "2025-03-18"),

    # ── MINOR ────────────────────────────────────────────────────────────────
    ("BUG-009",
     "Failed login shows raw error code instead of a user-friendly message",
     "Minor", "Autenticacion",
     "1. Go to /login.\n2. Enter a valid email and an incorrect password.\n3. Click Ingresar.",
     "Clear message displayed: 'Credenciales incorrectas. Por favor intenta de nuevo.'",
     "Generic error shown: 'These credentials do not match our records.' in English. No guidance or recovery option is provided.",
     "Chrome 124 / Windows 11", "Screenshot attached", "Open", "2025-03-22"),

    ("BUG-010",
     "Group Detail activity tab shows no loading state when saving attendance",
     "Minor", "Portal Profesor",
     "1. Go to /profesor/grupo/{id}.\n2. Select a date.\n3. Mark attendance for all students.\n4. Click Guardar Asistencia.",
     "Button shows a loading/disabled state while the request is in progress.",
     "Button remains active and clickable during submission. Clicking multiple times submits duplicate attendance records.",
     "Chrome 124 / Windows 11", "Loom link attached", "Open", "2025-04-02"),

    ("BUG-011",
     "Student grades page shows NaN% average when no activities have been evaluated",
     "Minor", "Calificaciones",
     "1. Log in as a student enrolled in a program with no evaluated activities.\n2. Go to /estudiante/calificaciones.",
     "Average grade shows '0%' or 'Sin evaluar' when no evaluations exist.",
     "Page displays 'NaN%' in the average grade card. This confuses students and looks unprofessional.",
     "Chrome 124 / Windows 11", "Screenshot attached", "Open", "2025-04-08"),
]

sev_style = {
    "Critical": (RED_FILL,    WHITE),
    "Major":    (ORANGE_FILL, WHITE),
    "Minor":    (YELLOW_FILL, "000000"),
}

CENTER_COLS = {1, 3, 8, 9, 10, 11}

for r_idx, bug in enumerate(bugs, 2):
    row_bg = GRAY_LIGHT if r_idx % 2 == 0 else WHITE
    ws.row_dimensions[r_idx].height = 90

    for c_idx, value in enumerate(bug, 1):
        c = ws.cell(row=r_idx, column=c_idx, value=value)
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if c_idx in CENTER_COLS else "left")
        c.border = brd

        if c_idx == 3:
            bg, fg = sev_style.get(value, (WHITE, "000000"))
            c.fill = solid(bg)
            c.font = Font(name="Arial", size=10, bold=True, color=fg)
        elif c_idx == 10:
            c.fill = solid(GREEN_FILL)
            c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        else:
            c.fill = solid(row_bg)
            c.font = Font(name="Arial", size=10)

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:K12"

# ── SHEET 2: Summary ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 26

def sc(row, col, value, bold=False, bg=None, fg="000000", size=11, halign="left"):
    c = ws2.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal=halign, vertical="center")
    c.border    = brd
    if bg:
        c.fill = solid(bg)
    return c

ws2.merge_cells("A1:B1")
t = ws2.cell(row=1, column=1, value="KAIROS — QA Bug Report Summary")
t.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
t.fill      = solid(BLUE_DARK)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

meta = [
    ("Tester",           "Marvin Eduardo Santos Perez"),
    ("Date",             "2025-04-08"),
    ("Platform",         "KAIROS — Academic Program Management"),
    ("Modules Tested",   "Matricula, Pagos, Autenticacion, Asistencia, Usuarios, Calificaciones, Portal Profesor"),
]
for i, (k, v) in enumerate(meta, 3):
    sc(i, 1, k, bold=True, bg=GRAY_LIGHT)
    sc(i, 2, v)
    ws2.row_dimensions[i].height = 22

ws2.row_dimensions[8].height = 14

sc(9, 1, "Severity",    bold=True, bg=BLUE_DARK, fg=WHITE, halign="center")
sc(9, 2, "Count",       bold=True, bg=BLUE_DARK, fg=WHITE, halign="center")
ws2.row_dimensions[9].height = 24

findings = [
    ("Critical",   "3",             RED_FILL,    WHITE),
    ("Major",      "5",             ORANGE_FILL, WHITE),
    ("Minor",      "3",             YELLOW_FILL, "000000"),
    ("Total Bugs", "=SUM(B10:B12)", BLUE_DARK,   WHITE),
]
for i, (label, val, bg, fg) in enumerate(findings, 10):
    sc(i, 1, label, bold=(label == "Total Bugs"), bg=bg, fg=fg, halign="center")
    sc(i, 2, val,   bold=True,                    bg=bg, fg=fg, halign="center")
    ws2.row_dimensions[i].height = 22

ws2.row_dimensions[14].height = 14

sc(15, 1, "Launch Readiness",  bold=True, bg=BLUE_DARK,   fg=WHITE, size=11, halign="center")
sc(15, 2, "Ready with Fixes",  bold=True, bg=ORANGE_FILL, fg=WHITE, size=11, halign="center")
ws2.row_dimensions[15].height = 28

wb.save(r"C:\Users\Lemon Lab Electronic\Desktop\sample-bug-report.xlsx")
print("OK")
